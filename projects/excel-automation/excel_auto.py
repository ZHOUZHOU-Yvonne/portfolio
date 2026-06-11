"""
Excel Automation Toolkit
Automate Excel workflows - merge, split, format, and generate reports.
"""

import os
import csv
from typing import List, Dict, Any, Optional
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

try:
    import pandas as pd
except ImportError:
    pd = None


class ReportGenerator:
    """Generate formatted Excel reports from data templates."""

    def __init__(self, template_path: Optional[str] = None):
        self.template = template_path
        if template_path and openpyxl:
            self.wb = openpyxl.load_workbook(template_path)
        elif openpyxl:
            self.wb = openpyxl.Workbook()
        else:
            self.wb = None
        self.ws = self.wb.active if self.wb else None
        self._style_header()

    def _style_header(self):
        if not self.ws:
            return
        header_font = Font(bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for cell in self.ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    def fill_data(self, data, start_row=2):
        """Fill worksheet with data from DataFrame or list of dicts."""
        if pd and isinstance(data, pd.DataFrame):
            data = data.values.tolist()
            headers = data.columns.tolist() if hasattr(data, 'columns') else []
        elif isinstance(data, list) and data and isinstance(data[0], dict):
            headers = list(data[0].keys())
            data = [[row.get(k, '') for k in headers] for row in data]
        else:
            headers = []
        if self.ws:
            for i, row in enumerate(data):
                for j, val in enumerate(row):
                    self.ws.cell(row=start_row + i, column=j + 1, value=val)

    def add_charts(self, chart_type='bar'):
        """Add chart to the worksheet."""
        if not self.ws:
            return
        chart = BarChart()
        chart.title = "Data Overview"
        chart.y_axis.title = "Values"
        chart.x_axis.title = "Categories"
        data = Reference(self.ws, min_col=2, min_row=1, max_row=self.ws.max_row, max_col=2)
        cats = Reference(self.ws, min_col=1, min_row=2, max_row=self.ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        self.ws.add_chart(chart, f"E{2}")

    def save(self, path: str):
        """Save the workbook."""
        if self.wb:
            self.wb.save(path)

    def auto_fit_columns(self):
        """Auto-fit column widths."""
        if not self.ws:
            return
        for col in self.ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            self.ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


class DataMerger:
    """Merge multiple Excel/CSV files into one."""

    @staticmethod
    def merge_excel(files: List[str], output: str, key_column: Optional[str] = None):
        """Merge multiple Excel files. Optionally join on key_column."""
        if not pd:
            raise ImportError("pandas required for merge operations")
        dfs = [pd.read_excel(f) for f in files]
        result = pd.concat(dfs, ignore_index=True)
        if key_column and key_column in result.columns:
            result = result.drop_duplicates(subset=[key_column])
        result.to_excel(output, index=False)
        return result

    @staticmethod
    def merge_csv(files: List[str], output: str):
        """Merge multiple CSV files."""
        if not pd:
            raise ImportError("pandas required for merge operations")
        dfs = [pd.read_csv(f) for f in files]
        result = pd.concat(dfs, ignore_index=True)
        result.to_csv(output, index=False)
        return result


class DataSplitter:
    """Split large Excel files by column values."""

    @staticmethod
    def split_by_column(file: str, column: str, output_dir: str):
        """Split Excel file into multiple files by unique column values."""
        if not pd:
            raise ImportError("pandas required for split operations")
        df = pd.read_excel(file)
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        files = []
        for val in df[column].unique():
            subset = df[df[column] == val]
            safe_name = str(val).replace('/', '_').replace('\\', '_')
            path = os.path.join(output_dir, f'{safe_name}.xlsx')
            subset.to_excel(path, index=False)
            files.append(path)
        return files

    @staticmethod
    def split_by_rows(file: str, rows_per_file: int, output_dir: str):
        """Split Excel file into chunks of N rows each."""
        if not pd:
            raise ImportError("pandas required for split operations")
        df = pd.read_excel(file)
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        files = []
        for i in range(0, len(df), rows_per_file):
            chunk = df.iloc[i:i + rows_per_file]
            path = os.path.join(output_dir, f'chunk_{i // rows_per_file + 1}.xlsx')
            chunk.to_excel(path, index=False)
            files.append(path)
        return files


class FormatCleaner:
    """Clean and normalize Excel formatting."""

    @staticmethod
    def clean_numbers(file: str, output: str, columns: Optional[List[str]] = None):
        """Convert text numbers to actual numbers, remove symbols."""
        if not pd:
            raise ImportError("pandas required for format operations")
        df = pd.read_excel(file)
        targets = columns or df.columns
        for col in targets:
            if col in df.columns:
                df[col] = df[col].astype(str).str.replace(r'[¥,$,€,,\s]', '', regex=True)
                df[col] = pd.to_numeric(df[col], errors='coerce')
        df.to_excel(output, index=False)
        return df

    @staticmethod
    def normalize_dates(file: str, output: str, date_columns: List[str], fmt: str = '%Y-%m-%d'):
        """Normalize date formats across columns."""
        if not pd:
            raise ImportError("pandas required for format operations")
        df = pd.read_excel(file)
        for col in date_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime(fmt)
        df.to_excel(output, index=False)
        return df

    @staticmethod
    def remove_duplicates(file: str, output: str, subset: Optional[List[str]] = None):
        """Remove duplicate rows from Excel file."""
        if not pd:
            raise ImportError("pandas required for format operations")
        df = pd.read_excel(file)
        before = len(df)
        df = df.drop_duplicates(subset=subset)
        after = len(df)
        df.to_excel(output, index=False)
        return {'before': before, 'after': after, 'removed': before - after}


if __name__ == '__main__':
    print("Excel Automation Toolkit")
    print("Usage: from excel_auto import ReportGenerator, DataMerger, DataSplitter, FormatCleaner")
